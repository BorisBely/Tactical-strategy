#!/usr/bin/env python3
"""Simple combat balance Excel: weapons as columns, distance/shot as rows, with charts."""

from __future__ import annotations

import math
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from combat_attachment_model import (
    REFERENCE_WEAPONS,
    combined_accuracy_quality,
    combined_aim_time_seconds,
    combined_recoil_control_quality,
    describe_sweet_spot,
    load_attachments,
    mod_affects_accuracy,
    mod_affects_aim,
    mod_affects_recoil,
)

ROOT = Path(__file__).resolve().parents[1]
SHOOTING = ROOT / "Assets" / "GameData" / "Shooting"
OUTPUT = ROOT / "Tools" / "CombatBalanceParameters.xlsx"

DISTANCES = list(range(0, 501, 50))
RECOIL_SHOTS = list(range(1, 11))
RECOIL_GRAPH_RECOVERY = 0.45

# Excel-only balance layer. Source assets are not changed.
# Accuracy is intentionally compressed: assault rifles stay within ~15%,
# support weapons sit ~20-30% below assault rifles, DMRs ~15-20% above them.
ASSAULT_ACCURACY_FACTORS = {
    "Weapon_AK47": 0.98,
    "Weapon_AK47_1": 0.99,
    "Weapon_AK47MOD1": 1.00,
    "Weapon_AK47S": 0.98,
    "Weapon_AK74": 1.00,
    "Weapon_AK74MOD1": 1.01,
    "Weapon_AK74U": 0.98,
    "Weapon_AK74UMOD1": 0.99,
    "Weapon_M4_ModA_1": 1.02,
    "Weapon_M4_ModA_2": 1.03,
    "Weapon_M16A_ModA_1": 1.04,
    "Weapon_MK18": 0.98,
}

ACCURACY_FACTORS = {
    **ASSAULT_ACCURACY_FACTORS,
    "Weapon_RPK47": 0.74,
    "Weapon_RPK47MOD1": 0.76,
    "Weapon_RPK74": 0.80,
    "Weapon_RPK74MOD1": 0.82,
    "Weapon_M16A4_ModA_2": 1.13,
    "Weapon_MK12": 1.16,
    "Weapon_SVD": 1.12,
    "Weapon_Sniper762x51": 1.18,
    "Weapon_Mosin": 1.15,
    "Weapon_M249": 0.78,
    "Weapon_PKM": 0.72,
    "Weapon_M2Browning_127": 0.50,
    "Weapon_MK19": 0.42,
}

ACCURACY_REFERENCE_BY_DISTANCE: dict[float, float] = {}

# role: (fade_in_start, plateau_start, plateau_end, fade_out_end, bonus)
ACCURACY_PROFILE_PLATEAUS = {
    "CqbShort": (0, 0, 25, 45, 0.055),
    "CqbControlled": (0, 5, 35, 55, 0.045),
    "ShotgunCqb": (0, 0, 15, 40, 0.040),
    "BattleRifle762Default": (5, 15, 45, 65, 0.035),
    "BattleRifle762WoodHandguard": (10, 20, 55, 75, 0.040),
    "BattleRifle762Mod1": (15, 25, 60, 80, 0.045),
    "Intermediate545": (10, 20, 55, 75, 0.040),
    "CarbineModA1": (15, 25, 60, 80, 0.045),
    "CarbineModA2": (20, 30, 65, 85, 0.050),
    "MidRifle": (25, 35, 75, 95, 0.050),
    "Marksman": (35, 45, 85, 100, 0.045),
    "Dmr": (45, 60, 100, 100, 0.040),
    "Support762": (25, 35, 75, 95, 0.045),
    "Support545": (30, 40, 80, 100, 0.045),
    "HeavySupport": (20, 30, 70, 90, 0.030),
    "GrenadeSupport": (15, 25, 60, 80, 0.025),
}

WEAPON_ORDER = [
    "Weapon_AK47",
    "Weapon_AK47_1",
    "Weapon_AK47MOD1",
    "Weapon_AK47S",
    "Weapon_AK74",
    "Weapon_AK74MOD1",
    "Weapon_AK74U",
    "Weapon_AK74UMOD1",
    "Weapon_RPK47",
    "Weapon_RPK47MOD1",
    "Weapon_RPK74",
    "Weapon_RPK74MOD1",
    "Weapon_M4_ModA_1",
    "Weapon_M4_ModA_2",
    "Weapon_M16A_ModA_1",
    "Weapon_M16A4_ModA_2",
    "Weapon_MK12",
    "Weapon_MK18",
    "Weapon_BenelliM4",
    "Weapon_Mosin",
    "Weapon_SVD",
    "Weapon_Sniper762x51",
    "Weapon_M249",
    "Weapon_PKM",
    "Weapon_M2Browning_127",
    "Weapon_MK19",
]

WEAPON_ROLE = {
    "Weapon_AK47": "BattleRifle762Default",
    "Weapon_AK47_1": "BattleRifle762WoodHandguard",
    "Weapon_AK47MOD1": "BattleRifle762Mod1",
    "Weapon_AK47S": "CqbControlled",
    "Weapon_AK74": "Intermediate545",
    "Weapon_AK74MOD1": "Intermediate545",
    "Weapon_AK74U": "CqbShort",
    "Weapon_AK74UMOD1": "CqbControlled",
    "Weapon_RPK47": "Support762",
    "Weapon_RPK47MOD1": "Support762",
    "Weapon_RPK74": "Support545",
    "Weapon_RPK74MOD1": "Support545",
    "Weapon_M4_ModA_1": "CarbineModA1",
    "Weapon_M4_ModA_2": "CarbineModA2",
    "Weapon_M16A_ModA_1": "MidRifle",
    "Weapon_M16A4_ModA_2": "Marksman",
    "Weapon_MK12": "Dmr",
    "Weapon_MK18": "CqbShort",
    "Weapon_BenelliM4": "ShotgunCqb",
    "Weapon_Mosin": "Dmr",
    "Weapon_SVD": "Marksman",
    "Weapon_Sniper762x51": "Dmr",
    "Weapon_M249": "Support545",
    "Weapon_PKM": "Support762",
    "Weapon_M2Browning_127": "HeavySupport",
    "Weapon_MK19": "GrenadeSupport",
}

ROLE_AIM = {
    "CqbShort": [(0, 0.92), (25, 1.08), (50, 2.55), (75, 4.15), (100, 5.85)],
    "CqbControlled": [(0, 0.84), (25, 1.10), (50, 2.36), (75, 3.79), (100, 5.33)],
    "ShotgunCqb": [(0, 1.05), (15, 1.18), (25, 1.45), (40, 1.95), (60, 2.80), (100, 4.20), (250, 6.20), (500, 8.50)],
    "CarbineModA1": [(0, 0.87), (25, 1.13), (50, 2.05), (75, 3.17), (100, 4.34)],
    "CarbineModA2": [(0, 0.90), (25, 1.12), (50, 1.98), (75, 3.05), (100, 4.20)],
    "BattleRifle762Default": [(0, 0.96), (25, 1.38), (50, 2.52), (75, 3.80), (100, 5.06)],
    "BattleRifle762WoodHandguard": [(0, 0.98), (25, 1.34), (50, 2.38), (75, 3.58), (100, 4.78)],
    "BattleRifle762Mod1": [(0, 1.02), (25, 1.36), (50, 2.34), (75, 3.48), (100, 4.62)],
    "Intermediate545": [(0, 0.90), (25, 1.25), (50, 2.16), (75, 3.22), (100, 4.29)],
    "MidRifle": [(0, 1.25), (25, 1.12), (50, 1.62), (75, 2.24), (100, 2.99)],
    "Marksman": [(0, 1.50), (25, 1.30), (50, 1.64), (75, 1.91), (100, 2.47)],
    "Dmr": [(0, 1.80), (25, 1.60), (50, 1.74), (75, 1.65), (100, 1.84)],
    "Support762": [(0, 1.55), (25, 1.35), (50, 1.69), (75, 2.00), (100, 2.59)],
    "Support545": [(0, 1.50), (25, 1.28), (50, 1.61), (75, 1.86), (100, 2.37)],
    "HeavySupport": [(0, 1.80), (25, 1.55), (50, 1.90), (75, 2.30), (100, 3.10)],
    "GrenadeSupport": [(0, 2.20), (25, 1.90), (50, 2.40), (75, 3.00), (100, 4.20)],
}

ROLE_DISP = {
    "CqbShort": [(0, 0.58), (25, 0.78), (50, 1.75), (75, 3.25), (100, 5.00)],
    "CqbControlled": [(0, 0.62), (25, 0.82), (50, 1.55), (75, 2.80), (100, 4.30)],
    "ShotgunCqb": [(0, 0.72), (15, 0.95), (25, 1.45), (40, 2.40), (60, 3.90), (100, 6.00), (250, 9.00), (500, 12.00)],
    "CarbineModA1": [(0, 0.73), (25, 0.86), (50, 1.08), (75, 1.64), (100, 2.42)],
    "CarbineModA2": [(0, 0.75), (25, 0.84), (50, 1.03), (75, 1.57), (100, 2.35)],
    "BattleRifle762Default": [(0, 0.80), (25, 0.98), (50, 1.32), (75, 2.12), (100, 3.18)],
    "BattleRifle762WoodHandguard": [(0, 0.79), (25, 0.94), (50, 1.22), (75, 1.92), (100, 2.88)],
    "BattleRifle762Mod1": [(0, 0.82), (25, 0.93), (50, 1.18), (75, 1.82), (100, 2.72)],
    "Intermediate545": [(0, 0.74), (25, 0.88), (50, 1.10), (75, 1.65), (100, 2.35)],
    "MidRifle": [(0, 0.90), (25, 0.75), (50, 0.65), (75, 1.00), (100, 1.70)],
    "Marksman": [(0, 1.00), (25, 0.82), (50, 0.58), (75, 0.70), (100, 1.20)],
    "Dmr": [(0, 1.15), (25, 1.00), (50, 0.70), (75, 0.50), (100, 0.62)],
    "Support762": [(0, 1.05), (25, 0.90), (50, 0.74), (75, 0.82), (100, 1.30)],
    "Support545": [(0, 1.00), (25, 0.85), (50, 0.66), (75, 0.70), (100, 1.05)],
    "HeavySupport": [(0, 1.40), (25, 1.20), (50, 1.05), (75, 1.15), (100, 1.80)],
    "GrenadeSupport": [(0, 1.80), (25, 1.50), (50, 1.35), (75, 1.55), (100, 2.40)],
}

ROLE_BURST = {
    "CqbShort": [(1, 1.00), (3, 1.50), (6, 3.10), (10, 6.00)],
    "CqbControlled": [(1, 1.00), (3, 1.42), (6, 2.75), (10, 5.20)],
    "ShotgunCqb": [(1, 1.00), (3, 1.65), (6, 3.40), (10, 6.50)],
    "CarbineModA1": [(1, 1.00), (3, 1.24), (6, 1.84), (10, 3.08)],
    "CarbineModA2": [(1, 1.00), (3, 1.23), (6, 1.82), (10, 3.02)],
    "BattleRifle762Default": [(1, 1.00), (3, 1.49), (6, 2.72), (10, 4.62)],
    "BattleRifle762WoodHandguard": [(1, 1.00), (3, 1.42), (6, 2.48), (10, 4.12)],
    "BattleRifle762Mod1": [(1, 1.00), (3, 1.40), (6, 2.38), (10, 3.92)],
    "Intermediate545": [(1, 1.00), (3, 1.30), (6, 2.10), (10, 3.50)],
    "MidRifle": [(1, 1.00), (3, 1.15), (6, 1.65), (10, 2.60)],
    "Marksman": [(1, 1.00), (3, 1.10), (6, 1.45), (10, 2.20)],
    "Dmr": [(1, 1.00), (3, 1.08), (6, 1.32), (10, 1.90)],
    "Support762": [(1, 1.00), (3, 1.18), (6, 1.55), (10, 2.50)],
    "Support545": [(1, 1.00), (3, 1.12), (6, 1.42), (10, 2.20)],
    "HeavySupport": [(1, 1.00), (3, 1.25), (6, 1.80), (10, 3.20)],
    "GrenadeSupport": [(1, 1.00), (3, 1.40), (6, 2.20), (10, 4.00)],
}

WEAPON_LABEL = {
    "Weapon_AK47": "AK-47",
    "Weapon_AK47_1": "AK-47 (wood)",
    "Weapon_AK47MOD1": "AK-47 MOD1",
    "Weapon_AK47S": "AK-47S",
    "Weapon_AK74": "AK-74",
    "Weapon_AK74MOD1": "AK-74 MOD1",
    "Weapon_AK74U": "AK-74U",
    "Weapon_AK74UMOD1": "AK-74U MOD1",
    "Weapon_RPK47": "RPK-47",
    "Weapon_RPK47MOD1": "RPK-47 MOD1",
    "Weapon_RPK74": "RPK-74",
    "Weapon_RPK74MOD1": "RPK-74 MOD1",
    "Weapon_M4_ModA_1": "M4 ModA1",
    "Weapon_M4_ModA_2": "M4 ModA2",
    "Weapon_M16A_ModA_1": "M16A ModA1",
    "Weapon_M16A4_ModA_2": "M16A4 ModA2",
    "Weapon_MK12": "MK12",
    "Weapon_MK18": "MK18",
    "Weapon_BenelliM4": "Benelli M4",
    "Weapon_Mosin": "Mosin",
    "Weapon_SVD": "SVD",
    "Weapon_Sniper762x51": "Sniper 7.62x51",
    "Weapon_M249": "M249",
    "Weapon_PKM": "PKM",
    "Weapon_M2Browning_127": "M2 Browning 12.7",
    "Weapon_MK19": "MK19",
}


def lerp_curve(keys: list[tuple[float, float]], x: float) -> float:
    if not keys:
        return 1.0
    if x <= keys[0][0]:
        return keys[0][1]
    if x >= keys[-1][0]:
        return keys[-1][1]
    for i in range(len(keys) - 1):
        x0, y0 = keys[i]
        x1, y1 = keys[i + 1]
        if x0 <= x <= x1:
            t = (x - x0) / (x1 - x0) if x1 != x0 else 0.0
            return y0 + t * (y1 - y0)
    return keys[-1][1]


def parse_curve_block(content: str, field_name: str) -> list[tuple[float, float]]:
    pattern = rf"{re.escape(field_name)}:\n(?:.*\n)*?      m_Curve:\n((?:      - serializedVersion: 3\n(?:        .+\n)+)+)"
    match = re.search(pattern, content)
    if not match:
        return []
    points: list[tuple[float, float]] = []
    for point in re.finditer(r"time: ([0-9.+-]+)\n        value: ([0-9.+-]+)", match.group(1)):
        points.append((float(point.group(1)), float(point.group(2))))
    return sorted(points)


def parse_weapon(path: Path) -> dict:
    content = path.read_text(encoding="utf-8")

    def get_float(name: str, default: float = 0.0) -> float:
        match = re.search(rf"  {name}: ([0-9.]+)", content)
        return float(match.group(1)) if match else default

    name = path.stem
    return {
        "name": name,
        "label": WEAPON_LABEL.get(name, name.replace("Weapon_", "")),
        "role": WEAPON_ROLE.get(name, "CarbineModA1"),
        "aim_base": get_float("m_AimTimeSeconds"),
        "disp_base": get_float("m_BaseShotDispersion"),
        "recoil_per_shot": get_float("m_RecoilPerShot", 1.0),
        "vertical_recoil": get_float("m_VerticalRecoil", 0.0),
        "horizontal_recoil": get_float("m_HorizontalRecoil", 0.0),
        "pattern_seed": get_float("m_RecoilPatternSeed", 0.0),
        "semi_recoil_mult": get_float("m_SemiAutoRecoilMultiplier", 0.85),
        "auto_recoil_mult": get_float("m_AutoRecoilMultiplier", 1.25),
        "recovery": get_float("m_RecoilRecoveryPerSecond", 0.7),
        "rpm": get_float("m_FireRateRpm", 600.0),
        "disp_curve": parse_curve_block(content, "m_DispersionMultiplierByDistance"),
        "aim_curve": parse_curve_block(content, "m_AimTimeMultiplierByDistance"),
        "burst_curve": parse_curve_block(content, "m_AutoBurstSpreadMultiplierByShot"),
    }


def curve_or_role(weapon: dict, curve_key: str, role_table: dict[str, list[tuple[float, float]]], x: float) -> float:
    curve = weapon.get(curve_key) or []
    if curve:
        return lerp_curve(curve, x)
    role = weapon["role"]
    return lerp_curve(role_table[role], x)


def accuracy_quality(weapon: dict, distance: float) -> float:
    mult = max(0.01, curve_or_role(weapon, "disp_curve", ROLE_DISP, distance))
    return 1.0 / mult


def profile_plateau_factor(weapon: dict, distance: float) -> float:
    plateau = ACCURACY_PROFILE_PLATEAUS.get(weapon["role"])
    if plateau is None:
        return 1.0

    fade_in_start, plateau_start, plateau_end, fade_out_end, bonus = plateau
    if distance < fade_in_start or distance > fade_out_end:
        return 1.0
    if plateau_start <= distance <= plateau_end:
        return 1.0 + bonus
    if distance < plateau_start:
        span = max(0.01, plateau_start - fade_in_start)
        return 1.0 + bonus * ((distance - fade_in_start) / span)

    span = max(0.01, fade_out_end - plateau_end)
    return 1.0 + bonus * (1.0 - ((distance - plateau_end) / span))


def adjusted_accuracy_quality(weapon: dict, distance: float) -> float:
    reference = ACCURACY_REFERENCE_BY_DISTANCE.get(distance)
    if reference is None:
        reference = accuracy_quality(weapon, distance)
    factor = ACCURACY_FACTORS.get(weapon["name"], 1.0)
    return reference * factor * profile_plateau_factor(weapon, distance)


def build_accuracy_reference(weapons: list[dict]) -> None:
    assault_weapons = [weapon for weapon in weapons if weapon["name"] in ASSAULT_ACCURACY_FACTORS]
    if not assault_weapons:
        return

    ACCURACY_REFERENCE_BY_DISTANCE.clear()
    for distance in DISTANCES:
        values = [accuracy_quality(weapon, distance) for weapon in assault_weapons]
        ACCURACY_REFERENCE_BY_DISTANCE[distance] = sum(values) / len(values)


def aim_time_seconds(weapon: dict, distance: float) -> float:
    mult = max(0.01, curve_or_role(weapon, "aim_curve", ROLE_AIM, distance))
    return max(0.01, weapon["aim_base"] * mult)


def burst_mult(weapon: dict, shot: int) -> float:
    curve = weapon.get("burst_curve") or []
    if curve:
        return max(0.01, lerp_curve(curve, float(shot)))
    return max(0.01, lerp_curve(ROLE_BURST[weapon["role"]], float(shot)))


def recoil_control_quality(weapon: dict, shot: int) -> float:
    base = max(0.01, weapon["recoil_per_shot"])
    recoil_added = base * weapon["auto_recoil_mult"]
    if recoil_added <= 0:
        return 1.0

    fire_interval = 60.0 / max(1.0, weapon["rpm"])
    recovery_per_shot = weapon["recovery"] * RECOIL_GRAPH_RECOVERY * fire_interval
    accumulated = 0.0
    for _ in range(1, shot):
        accumulated += recoil_added
        accumulated = max(0.0, accumulated - recovery_per_shot)

    burden = (1.0 + accumulated) * burst_mult(weapon, shot)
    return (1.0 / base) / max(0.01, burden)


def load_weapons() -> list[dict]:
    weapons: list[dict] = []
    for asset_name in WEAPON_ORDER:
        matches = list(SHOOTING.rglob(f"{asset_name}.asset"))
        if not matches:
            print(f"Warning: missing {asset_name}", file=sys.stderr)
            continue
        weapons.append(parse_weapon(matches[0]))
    return weapons


def style_header_row(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize_columns(ws, max_width: int = 14) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, 10), max_width)
    ws.column_dimensions["A"].width = 16


def write_mod_loadout_sheet(
    wb: Workbook,
    title: str,
    row_label: str,
    row_values: list,
    weapon: dict,
    attachments: list[dict],
    value_fn,
    chart_y_title: str,
    number_format: str = "0.000",
    subtitle: str = "",
) -> None:
    ws = wb.create_sheet(title)
    if subtitle:
        ws.append([subtitle])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(attachments))
        ws["A1"].font = Font(bold=True, size=11)

    header_row = 2 if subtitle else 1
    data_start_row = header_row + 1
    ws.cell(row=header_row, column=1, value=row_label)
    ws.cell(row=header_row, column=2, value=f"База ({weapon['label']})")
    for col_idx, attachment in enumerate(attachments, start=3):
        ws.cell(row=header_row, column=col_idx, value=f"+ {attachment['label']}")
    style_header_row(ws, row=header_row)

    for row_offset, row_value in enumerate(row_values):
        row_idx = data_start_row + row_offset
        ws.cell(row=row_idx, column=1, value=row_value)
        ws.cell(row=row_idx, column=2, value=round(value_fn(weapon, [], row_value), 3))
        for col_idx, attachment in enumerate(attachments, start=3):
            ws.cell(row=row_idx, column=col_idx, value=round(value_fn(weapon, [attachment], row_value), 3))

    last_row = data_start_row + len(row_values) - 1
    last_col = 2 + len(attachments)
    for row_idx in range(data_start_row, last_row + 1):
        for col_idx in range(2, last_col + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = number_format

    autosize_columns(ws, max_width=16)

    chart = LineChart()
    chart.title = title
    chart.y_axis.title = chart_y_title
    chart.x_axis.title = row_label
    chart.height = 12
    chart.width = 24
    chart.legend.position = "r"

    data_ref = Reference(ws, min_col=2, max_col=last_col, min_row=header_row, max_row=last_row)
    cats_ref = Reference(ws, min_col=1, min_row=data_start_row, max_row=last_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    anchor_row = last_row + 4
    ws.add_chart(chart, f"A{anchor_row}")


def write_attachment_params_sheet(
    wb: Workbook,
    weapons_by_name: dict[str, dict],
    distances: list[float],
    recoil_shots: list[int],
) -> None:
    ws = wb.create_sheet("Моды")
    headers = [
        "Модуль",
        "Платформа",
        "Тип",
        "Слот",
        "Aim×",
        "Recoil×",
        "Semi×",
        "Auto×",
        "Reload×",
        "Range×",
        "Disp curve",
        "Aim curve",
        "Sweet spot",
        "Δ точн.@25м",
        "Δ aim@25м",
        "Δ отдача@5",
    ]
    ws.append(headers)
    style_header_row(ws)

    attachments = load_attachments()
    for attachment in attachments:
        platform = attachment["platform"] or "?"
        ref_weapon_name = REFERENCE_WEAPONS.get(platform or "")
        ref_weapon = weapons_by_name.get(ref_weapon_name) if ref_weapon_name else None

        disp_curve_label = "library" if attachment["uses_rail_library"] else (
            "library" if attachment["uses_optic_library"] else (
                "custom" if not all(abs(v - 1.0) < 0.001 for _, v in attachment["disp_curve"]) else "flat"
            )
        )
        aim_curve_label = "library" if attachment["uses_optic_library"] else (
            "custom" if not all(abs(v - 1.0) < 0.001 for _, v in attachment["aim_curve"]) else "flat"
        )

        delta_accuracy = ""
        delta_aim = ""
        delta_recoil = ""
        if ref_weapon is not None:
            bare_accuracy = accuracy_quality(ref_weapon, 25)
            mod_accuracy = combined_accuracy_quality(ref_weapon, [attachment], 25, curve_or_role_disp)
            delta_accuracy = round(mod_accuracy - bare_accuracy, 3)

            bare_aim = aim_time_seconds(ref_weapon, 25)
            mod_aim = combined_aim_time_seconds(ref_weapon, [attachment], 25, curve_or_role_aim)
            delta_aim = round(mod_aim - bare_aim, 3)

            bare_recoil = recoil_control_quality(ref_weapon, 5)
            mod_recoil = combined_recoil_control_quality(ref_weapon, [attachment], 5, burst_mult)
            delta_recoil = round(mod_recoil - bare_recoil, 3)

        ws.append([
            attachment["label"],
            platform,
            attachment["type_label"],
            attachment["slot_label"],
            round(attachment["aim_time_modifier"], 3),
            round(attachment["recoil_modifier"], 3),
            round(attachment["semi_recoil_modifier"], 3),
            round(attachment["auto_recoil_modifier"], 3),
            round(attachment["reload_time_modifier"], 3),
            round(attachment["effective_range_modifier"], 3),
            disp_curve_label,
            aim_curve_label,
            describe_sweet_spot(attachment, distances),
            delta_accuracy,
            delta_aim,
            delta_recoil,
        ])

    for row_idx in range(2, 2 + len(attachments)):
        for col_idx in range(5, 11):
            ws.cell(row=row_idx, column=col_idx).number_format = "0.000"
        for col_idx in range(14, 17):
            ws.cell(row=row_idx, column=col_idx).number_format = "0.000"

    autosize_columns(ws, max_width=18)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["L"].width = 14


def curve_or_role_disp(weapon: dict, distance: float) -> float:
    return curve_or_role(weapon, "disp_curve", ROLE_DISP, distance)


def curve_or_role_aim(weapon: dict, distance: float) -> float:
    return curve_or_role(weapon, "aim_curve", ROLE_AIM, distance)


def write_platform_mod_sheets(
    wb: Workbook,
    platform: str,
    weapons_by_name: dict[str, dict],
) -> None:
    weapon_name = REFERENCE_WEAPONS[platform]
    weapon = weapons_by_name.get(weapon_name)
    if weapon is None:
        print(f"Warning: missing reference weapon {weapon_name} for {platform}", file=sys.stderr)
        return

    attachments = load_attachments(platform)
    accuracy_mods = [a for a in attachments if mod_affects_accuracy(a)]
    aim_mods = [a for a in attachments if mod_affects_aim(a)]
    recoil_mods = [a for a in attachments if mod_affects_recoil(a)]

    subtitle = f"Эталон: {weapon['label']}. Каждый столбец = базовое оружие + один модуль."

    if accuracy_mods:
        write_mod_loadout_sheet(
            wb,
            f"{platform} точность",
            "Дист, м",
            DISTANCES,
            weapon,
            accuracy_mods,
            lambda w, atts, d: combined_accuracy_quality(w, atts, d, curve_or_role_disp),
            "Качество точности",
            subtitle=subtitle,
        )

    if aim_mods:
        write_mod_loadout_sheet(
            wb,
            f"{platform} прицел",
            "Дист, м",
            DISTANCES,
            weapon,
            aim_mods,
            lambda w, atts, d: combined_aim_time_seconds(w, atts, d, curve_or_role_aim),
            "Время, сек",
            subtitle=subtitle,
        )

    if recoil_mods:
        write_mod_loadout_sheet(
            wb,
            f"{platform} отдача",
            "Выстрел",
            RECOIL_SHOTS,
            weapon,
            recoil_mods,
            lambda w, atts, s: combined_recoil_control_quality(w, atts, s, burst_mult),
            "Контроль отдачи",
            subtitle=subtitle,
        )


def write_matrix_sheet(
    wb: Workbook,
    title: str,
    row_label: str,
    row_values: list,
    weapons: list[dict],
    value_fn,
    chart_y_title: str,
    number_format: str = "0.00",
) -> None:
    ws = wb.create_sheet(title)
    ws.append([row_label] + [w["label"] for w in weapons])
    style_header_row(ws)

    for row_value in row_values:
        row = [row_value]
        for weapon in weapons:
            row.append(round(value_fn(weapon, row_value), 3))
        ws.append(row)

    for row_idx in range(2, 2 + len(row_values)):
        for col_idx in range(2, 2 + len(weapons)):
            ws.cell(row=row_idx, column=col_idx).number_format = number_format

    autosize_columns(ws)

    chart = LineChart()
    chart.title = title
    chart.y_axis.title = chart_y_title
    chart.x_axis.title = row_label
    chart.height = 12
    chart.width = 24
    chart.legend.position = "r"

    data_ref = Reference(ws, min_col=2, max_col=1 + len(weapons), min_row=1, max_row=1 + len(row_values))
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=1 + len(row_values))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    anchor_row = len(row_values) + 4
    ws.add_chart(chart, f"A{anchor_row}")


def write_description_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Описание"
    rows = [
        ["Combat Balance — упрощённая таблица", ""],
        ["", ""],
        ["Лист", "Что показывает"],
        ["Точность", "Качество точности из Unity assets (1 / dispersion multiplier). Без модулей."],
        ["Прицеливание", "Время полного прицеливания, сек. Меньше = быстрее. Без модулей."],
        ["Отдача (УСТАРЕЛО)", "Старая модель RecoilPerShot / P. Не использовать для калибровки offset-отдачи. Ждём новую таблицу."],
        ["Параметры", "Сырые поля каждого Weapon_*.asset, включая новые Vertical/Horizontal/Recovery °/с. Сюда вносить обновлённые данные."],
        ["Моды", "Плоские множители и дистанционные кривые каждого модуля. Range× = EffectiveRangeModifier. Δ — изменение vs эталонное оружие."],
        ["M4 точность / прицел / отдача", "M4 ModA1 + по одному модулю в столбце. График = влияние модов."],
        ["AK точность / прицел / отдача", "AK-74 + по одному модулю в столбце. График = влияние модов."],
        ["", ""],
        ["Строки", "0, 50, 100 … 500 м — для точности и прицеливания"],
        ["Строки отдачи", "Номер выстрела в очереди FullAuto (1–10)"],
        ["Столбцы", "Оружие в фиксированном порядке"],
        ["Графики", "Линейные графики под каждой таблицей"],
        ["Benelli M4", "Роль ShotgunCqb: 0–15 лучший, 15–25 сильный, 25–40 рабочий, 40–60 хуже АК, 60+ почти бесполезен. Эффективность дроби = паттерн + falloff, не только лист Точность."],
        ["M2 / MK19", "Турельные стволы добавлены 2026-08-21. Раньше в книге их не было."],
        ["", ""],
        ["Важно", "Точность и прицеливание читаются из Unity assets. Перезапекание: python Tools/bake_weapon_combat_balance.py"],
        ["Обновление", "python Tools/export_combat_balance_excel.py"],
    ]
    for row in rows:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=14)
    autosize_columns(ws, 70)


def write_weapon_params_sheet(wb: Workbook, weapons: list[dict]) -> None:
    ws = wb.create_sheet("Параметры")
    headers = [
        "Asset",
        "Название",
        "Роль",
        "RPM",
        "AimTime с",
        "BaseDispersion",
        "RecoilPerShot (legacy)",
        "VerticalRecoil °",
        "HorizontalRecoil °",
        "PatternSeed",
        "Recovery °/с",
        "SemiRecoil×",
        "AutoRecoil×",
        "Нужна новая калибровка",
    ]
    ws.append(headers)
    style_header_row(ws)
    for weapon in weapons:
        needs_new = "да — offset-модель, значения стартовые"
        ws.append([
            weapon["name"],
            weapon["label"],
            weapon["role"],
            weapon["rpm"],
            weapon["aim_base"],
            weapon["disp_base"],
            weapon["recoil_per_shot"],
            weapon.get("vertical_recoil", 0.0),
            weapon.get("horizontal_recoil", 0.0),
            weapon.get("pattern_seed", 0.0),
            weapon["recovery"],
            weapon.get("semi_recoil_mult", 0.0),
            weapon["auto_recoil_mult"],
            needs_new,
        ])
    autosize_columns(ws, 28)


def main() -> None:
    weapons = load_weapons()
    if not weapons:
        raise SystemExit("No weapon assets found.")
    build_accuracy_reference(weapons)
    weapons_by_name = {weapon["name"]: weapon for weapon in weapons}

    wb = Workbook()
    write_description_sheet(wb)
    write_weapon_params_sheet(wb, weapons)
    write_attachment_params_sheet(wb, weapons_by_name, DISTANCES, RECOIL_SHOTS)

    write_matrix_sheet(
        wb,
        "Точность",
        "Дист, м",
        DISTANCES,
        weapons,
        accuracy_quality,
        "Качество точности",
        "0.000",
    )
    write_matrix_sheet(
        wb,
        "Прицеливание",
        "Дист, м",
        DISTANCES,
        weapons,
        aim_time_seconds,
        "Время, сек",
        "0.000",
    )
    write_matrix_sheet(
        wb,
        "Отдача",
        "Выстрел",
        RECOIL_SHOTS,
        weapons,
        recoil_control_quality,
        "Контроль отдачи",
        "0.000",
    )

    write_platform_mod_sheets(wb, "M4", weapons_by_name)
    write_platform_mod_sheets(wb, "AK", weapons_by_name)
    write_platform_mod_sheets(wb, "SVD", weapons_by_name)
    write_platform_mod_sheets(wb, "Sniper", weapons_by_name)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")
    print(f"Weapons: {len(weapons)}")
    print(f"Attachments: {len(load_attachments())}")


if __name__ == "__main__":
    main()

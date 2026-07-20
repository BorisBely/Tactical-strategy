# -*- coding: utf-8 -*-
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

root = Path(r"d:\UnityProjects\My project 001")
weapons = sorted((root / "Assets" / "GameData" / "Shooting").rglob("Weapon_*.asset"))


def parse_float(text, key, default=None):
	m = re.search(rf"{key}:\s*([-+0-9.eE]+)", text)
	return float(m.group(1)) if m else default


def extract_curve_keys(block):
	if not block:
		return {}
	return {
		float(t): float(v)
		for t, v in re.findall(r"time:\s*([0-9.]+)\s*\n\s*value:\s*([-+0-9.eE]+)", block)
	}


def role_values(curve):
	out = []
	distances = (0.0, 125.0, 250.0, 375.0, 500.0)
	if not curve:
		return [None] * 5
	keys = sorted(curve.keys())
	for d in distances:
		if d in curve:
			out.append(curve[d])
			continue
		if d <= keys[0]:
			out.append(curve[keys[0]])
			continue
		if d >= keys[-1]:
			out.append(curve[keys[-1]])
			continue
		for i in range(len(keys) - 1):
			t0, t1 = keys[i], keys[i + 1]
			if t0 <= d <= t1:
				v0, v1 = curve[t0], curve[t1]
				t = 0.0 if t1 == t0 else (d - t0) / (t1 - t0)
				out.append(v0 + (v1 - v0) * t)
				break
	return out


def parse_weapon(path: Path):
	text = path.read_text(encoding="utf-8", errors="replace")
	name = path.stem.replace("Weapon_", "")
	aim_section = re.search(
		r"m_AimTimeMultiplierByDistance:.*?m_Curve:(.*?)(?:m_PreInfinity)",
		text,
		re.S,
	)
	disp_section = re.search(
		r"m_DispersionMultiplierByDistance:.*?m_Curve:(.*?)(?:m_PreInfinity)",
		text,
		re.S,
	)
	aim = extract_curve_keys(aim_section.group(1) if aim_section else "")
	disp = extract_curve_keys(disp_section.group(1) if disp_section else "")
	return {
		"name": name,
		"aimBase": parse_float(text, "m_AimTimeSeconds"),
		"dispersion": parse_float(text, "m_BaseShotDispersion"),
		"range": parse_float(text, "m_EffectiveRangeMeters"),
		"aimRole": role_values(aim),
		"dispRole": role_values(disp),
	}


rows = [parse_weapon(p) for p in weapons]

rl_path = root / "Assets" / "GameData" / "Combat" / "RocketLauncherData.asset"
rl = rl_path.read_text(encoding="utf-8")
parts = rl.split("m_DisposableDistanceAimProfile:")
rpg_part = parts[0]
dis_part = parts[1] if len(parts) > 1 else ""


def profile_curves(block):
	disp_m = re.search(
		r"m_DispersionMultiplierByDistance:.*?m_Curve:(.*?)m_PreInfinity",
		block,
		re.S,
	)
	aim_m = re.search(
		r"m_AimTimeMultiplierByDistance:.*?m_Curve:(.*?)m_PreInfinity",
		block,
		re.S,
	)
	return (
		extract_curve_keys(disp_m.group(1) if disp_m else ""),
		extract_curve_keys(aim_m.group(1) if aim_m else ""),
	)


rpg_disp, rpg_aim = profile_curves(rpg_part)
dis_disp, dis_aim = profile_curves(dis_part)

rockets = [
	{
		"name": "RPG-7",
		"aimBase": parse_float(rl, "m_RpgAimTimeSeconds"),
		"dispersionDeg": parse_float(rl, "m_RpgBaseDispersionDegrees"),
		"muzzle": parse_float(rl, "m_RpgMuzzleSpeed"),
		"aimRole": role_values(rpg_aim),
		"dispRole": role_values(rpg_disp),
	},
	{
		"name": "Disposable RL",
		"aimBase": parse_float(rl, "m_DisposableAimTimeSeconds"),
		"dispersionDeg": parse_float(rl, "m_DisposableBaseDispersionDegrees"),
		"muzzle": parse_float(rl, "m_DisposableMuzzleSpeed"),
		"aimRole": role_values(dis_aim),
		"dispRole": role_values(dis_disp),
	},
]

wb = Workbook()
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(color="FFFFFF", bold=True)

ws = wb.active
ws.title = "Aim Time"
ws.append(
	[
		"Weapon",
		"Type",
		"AimBaseSec",
		"Mul0",
		"Mul125",
		"Mul250",
		"Mul375",
		"Mul500",
		"Sec0",
		"Sec125",
		"Sec250",
		"Sec375",
		"Sec500",
		"Notes",
	]
)
for r in rows:
	base = r["aimBase"] or 0.0
	ar = r["aimRole"]
	secs = [round(base * (ar[i] or 0.0), 3) if ar[i] is not None else "" for i in range(5)]
	ws.append(
		[
			r["name"],
			"Firearm",
			base,
			*[round(x, 3) if x is not None else "" for x in ar],
			*secs,
			"WeaponDefinition asset",
		]
	)
for r in rockets:
	base = r["aimBase"] or 0.0
	ar = r["aimRole"]
	secs = []
	for i, d in enumerate((0.0, 125.0, 250.0, 375.0, 500.0)):
		raw = base * (ar[i] or 0.0)
		floor = 3.0 * (d / 500.0)
		secs.append(round(max(raw, floor), 3))
	ws.append(
		[
			r["name"],
			"RocketLauncher",
			base,
			*[round(x, 3) if x is not None else "" for x in ar],
			*secs,
			"RocketLauncherData; min aim floor 3s at 500m; rank WeaponHandling",
		]
	)

ws2 = wb.create_sheet("Accuracy")
ws2.append(
	[
		"Weapon",
		"Type",
		"BaseDispersion",
		"Unit",
		"Mul0",
		"Mul125",
		"Mul250",
		"Mul375",
		"Mul500",
		"Total0",
		"Total125",
		"Total250",
		"Total375",
		"Total500",
		"Notes",
	]
)
for r in rows:
	base = r["dispersion"] or 0.0
	dr = r["dispRole"]
	totals = [round(base * (dr[i] or 0.0), 3) if dr[i] is not None else "" for i in range(5)]
	ws2.append(
		[
			r["name"],
			"Firearm",
			base,
			"BaseShotDispersion (*0.35 -> half-angle deg)",
			*[round(x, 3) if x is not None else "" for x in dr],
			*totals,
			"WeaponDefinition asset",
		]
	)
for r in rockets:
	base = r["dispersionDeg"] or 0.0
	dr = r["dispRole"]
	totals = [round(base * (dr[i] or 0.0), 3) if dr[i] is not None else "" for i in range(5)]
	ws2.append(
		[
			r["name"],
			"RocketLauncher",
			base,
			"half-angle degrees (cone)",
			*[round(x, 3) if x is not None else "" for x in dr],
			*totals,
			f"muzzle {r['muzzle']} m/s; rank Marksmanship",
		]
	)

ws3 = wb.create_sheet("Coverage")
ws3.append(["AssetInProject", "InOldCombatBalanceTablesMd", "InThisExcel"])
detailed_md = {"AK47", "M4_ModA_2", "BenelliM4"}
for r in rows:
	ws3.append(
		[
			r["name"],
			"yes (detailed)" if r["name"] in detailed_md else "no",
			"yes",
		]
	)
ws3.append(["RPG-7", "no", "yes"])
ws3.append(["Disposable RL", "no", "yes"])

for sheet in wb.worksheets:
	for cell in sheet[1]:
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(wrap_text=True, vertical="center")
	for col in sheet.columns:
		letter = col[0].column_letter
		width = 14
		for cell in col[:40]:
			width = max(width, min(40, len(str(cell.value or "")) + 2))
		sheet.column_dimensions[letter].width = width

out = root / "Assets" / "Docs" / "CombatBalance" / "WeaponAimAccuracyBalance.xlsx"
out.parent.mkdir(parents=True, exist_ok=True)
wb.save(out)
print("saved", out)
print("firearms", len(rows))
for r in rows:
	print(f" - {r['name']}: aim={r['aimBase']} disp={r['dispersion']}")
print("rockets:")
for r in rockets:
	print(f" - {r['name']}: aim={r['aimBase']} deg={r['dispersionDeg']} muzzle={r['muzzle']}")
